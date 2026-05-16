import os
import pandas as pd
import re

def load_staff(filepath: str) -> pd.DataFrame:
    """
    Load staff data from the first sheet of the Excel workbook.
    """
    from openpyxl import load_workbook

    if not os.path.exists(filepath):
        print(f"Error: File not found at {filepath}")
        return pd.DataFrame()

    try:
        # Load workbook directly (no need for temp copy on Linux/Streamlit Cloud)
        wb = load_workbook(filepath, data_only=True)
        if not wb.worksheets:
            return pd.DataFrame()
        ws = wb.active if wb.active else wb.worksheets[0]
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return pd.DataFrame()

    # ── Find the header row ────────────────────────────────────────────────────
    header_row_num = None
    for row in ws.iter_rows():
        vals = [str(c.value).strip().lower() for c in row if c.value]
        joined = " ".join(vals)
        if "name of the faculty" in joined or ("name" in joined and "designation" in joined):
            header_row_num = row[0].row
            break

    if header_row_num is None:
        return pd.DataFrame()

    # ── Read header columns ───────────────────────────────────────────────────
    try:
        header_row = list(ws.iter_rows(min_row=header_row_num, max_row=header_row_num))[0]
    except (IndexError, Exception):
        return pd.DataFrame()

    col_map = {}
    for i, cell in enumerate(header_row):
        raw = str(cell.value or "").strip().lower().replace(".", "").replace(" ", "")
        col_map[raw] = i

    def _find_col(*keys):
        for k in keys:
            nk = k.lower().replace(".", "").replace(" ", "")
            if nk in col_map:
                return col_map[nk]
            for mapped_key, idx in col_map.items():
                if nk in mapped_key or mapped_key in nk:
                    return idx
        return None

    slno_idx   = _find_col("slno", "sl.no")
    ecode_idx  = _find_col("ecode", "e.code")
    name_idx   = _find_col("nameofthefaculty", "name of the faculty", "name")
    desig_idx  = _find_col("designation", "desig")
    mobile_idx = _find_col("mobileno", "mobile no", "mobile no.", "phone")

    if name_idx is None:
        return pd.DataFrame()

    def _cell_val(cells: list, idx) -> str:
        if idx is None or idx >= len(cells):
            return ""
        v = cells[idx].value
        if v is None:
            return ""
        return " ".join(str(v).replace("\n", " ").split()).strip()

    # ── Extract data rows ──────────────────────────────────────────────────────
    records = []
    for row in ws.iter_rows(min_row=header_row_num + 1):
        cells = list(row)
        name = _cell_val(cells, name_idx)
        if not name or name.lower() in ("nan", "none", "name of the faculty", "name", "faculty name"):
            continue

        records.append({
            "Sl.No":       _cell_val(cells, slno_idx),
            "E.Code":      _cell_val(cells, ecode_idx),
            "Name":        name,
            "Designation": _cell_val(cells, desig_idx),
            "Mobile No":   _cell_val(cells, mobile_idx),
        })

    return pd.DataFrame(records)


def load_invigilation(filepath: str) -> pd.DataFrame:
    """Load invigilation data using openpyxl."""
    from openpyxl import load_workbook

    if not os.path.exists(filepath):
        return pd.DataFrame()

    try:
        wb = load_workbook(filepath, data_only=True)
        if not wb.worksheets:
            return pd.DataFrame()
        ws = wb.active if wb.active else wb.worksheets[0]
    except Exception:
        return pd.DataFrame()

    # Find header row
    header_row_num = None
    for row in ws.iter_rows():
        vals = [str(c.value or "").strip().lower() for c in row if c.value]
        joined = " ".join(vals)
        if "room" in joined or "floor" in joined or "faculty" in joined or "invigilator" in joined:
            header_row_num = row[0].row
            break

    if header_row_num is None:
        return pd.DataFrame()

    try:
        header_cells = list(ws.iter_rows(min_row=header_row_num, max_row=header_row_num))[0]
        columns = [str(c.value or "").strip() or f"Col{i}" for i, c in enumerate(header_cells)]
    except Exception:
        return pd.DataFrame()

    records = []
    for row in ws.iter_rows(min_row=header_row_num + 1):
        vals = [" ".join(str(c.value or "").replace("\n", " ").split()) for c in row]
        if any(v for v in vals):
            records.append(dict(zip(columns, vals)))

    df = pd.DataFrame(records)
    return df[df.apply(lambda r: any(r.values), axis=1)].reset_index(drop=True)
